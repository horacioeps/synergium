from __future__ import annotations

import logging
from dataclasses import asdict, dataclass, field
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .analyze import Clasificacion
from .sources.base import Candidato

log = logging.getLogger(__name__)

COLUMNAS = [
    "fecha_deteccion", "plataforma", "autor_o_cuenta", "titulo_o_extracto",
    "senal_de_dolor", "URL", "tema", "pais", "categoria",
    "estado", "proximo_paso", "relevancia_1_5", "mensaje_sugerido",
    "fuente_id", "es_competidor", "tipo_actor",
]


@dataclass
class Registro:
    URL: str
    fecha_deteccion: str
    plataforma: str
    autor_o_cuenta: str
    titulo_o_extracto: str
    senal_de_dolor: str
    tema: str
    pais: str
    categoria: str          # 'personal' | 'INV_AMP'
    estado: str             # 'nuevo' | 'descartado' | 'contactado' | 'respondio' ...
    proximo_paso: str
    relevancia_1_5: int
    mensaje_sugerido: str
    fuente_id: str
    es_competidor: bool
    tipo_actor: str


def _categoria_por_defecto(objetivo: str | None) -> str:
    # Convencion v1: todo cae en INV_AMP salvo decision posterior del usuario
    return "INV_AMP"


def registro_de(c: Candidato, clf: Clasificacion, mensaje: str | None) -> Registro:
    titulo = (c.titulo or "").strip()
    extracto = (c.extracto or "").strip()
    titulo_o_extracto = titulo if titulo else extracto[:200]
    return Registro(
        URL=c.url,
        fecha_deteccion=c.fecha_deteccion.isoformat(),
        plataforma=c.plataforma or "",
        autor_o_cuenta=(c.autor or ""),
        titulo_o_extracto=titulo_o_extracto,
        senal_de_dolor=(clf.motivo or "").strip(),
        tema=clf.tema or "",
        pais=clf.pais_inferido or (c.pais or "desconocido"),
        categoria=_categoria_por_defecto(c.objetivo),
        estado="nuevo",
        proximo_paso=("contactar" if clf.relevancia >= 4 else "revisar"),
        relevancia_1_5=int(clf.relevancia),
        mensaje_sugerido=(mensaje or ""),
        fuente_id=c.fuente_id,
        es_competidor=bool(clf.es_competidor),
        tipo_actor=clf.tipo_actor or "otro",
    )


def cargar_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=COLUMNAS)
    return pd.read_csv(path)


def merge_incremental(csv_path: Path, nuevos: list[Registro]) -> tuple[pd.DataFrame, int]:
    df_old = cargar_csv(csv_path)
    df_new = pd.DataFrame([asdict(r) for r in nuevos], columns=COLUMNAS)

    if df_old.empty:
        df = df_new
        anadidos = len(df_new)
    else:
        existing_urls = set(df_old["URL"].astype(str).tolist())
        mask_realmente_nuevos = ~df_new["URL"].astype(str).isin(existing_urls)
        df_add = df_new[mask_realmente_nuevos]
        anadidos = len(df_add)
        df = pd.concat([df_old, df_add], ignore_index=True)

    # ordenar por relevancia desc luego fecha desc
    df = df.sort_values(
        by=["relevancia_1_5", "fecha_deteccion"], ascending=[False, False]
    ).reset_index(drop=True)
    return df, anadidos


def guardar(df: pd.DataFrame, csv_path: Path, md_path: Path,
            xlsx_path: Path | None = None) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(csv_path, index=False)
    _escribir_markdown(df, md_path)
    if xlsx_path is not None:
        _escribir_excel(df, xlsx_path)


# Colores por relevancia (ARGB sin #)
_FILL_REL = {
    5: "FF63BE7B",   # verde fuerte
    4: "FFB1D580",   # verde claro
    3: "FFFFEB84",   # amarillo
    2: "FFF8B26A",   # naranja claro
    1: "FFF8696B",   # rojo claro
}
_FILL_COMPETIDOR = "FFD9D9D9"   # gris para competidores (independiente del color anterior)
_HEADER_FILL = PatternFill("solid", fgColor="FF305496")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")

_ESTADOS = ["nuevo", "revisar", "contactar", "contactado", "respondio",
            "agendado", "descartado", "competidor"]

_COL_WIDTHS = {
    "relevancia_1_5": 6,
    "pais": 6,
    "tipo_actor": 14,
    "categoria": 10,
    "estado": 13,
    "proximo_paso": 14,
    "tema": 28,
    "titulo_o_extracto": 50,
    "senal_de_dolor": 45,
    "mensaje_sugerido": 55,
    "URL": 45,
    "fecha_deteccion": 12,
    "plataforma": 9,
    "autor_o_cuenta": 18,
    "fuente_id": 22,
    "es_competidor": 8,
}

_COL_ORDER_PROSPECTOS = [
    "relevancia_1_5", "estado", "pais", "tipo_actor", "tema",
    "titulo_o_extracto", "senal_de_dolor", "mensaje_sugerido",
    "URL", "proximo_paso", "categoria", "plataforma", "autor_o_cuenta",
    "fecha_deteccion", "fuente_id", "es_competidor",
]


def _hoja_tabla(wb: Workbook, titulo: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(titulo)
    if df.empty:
        ws.append(["(sin filas)"])
        return

    cols = [c for c in _COL_ORDER_PROSPECTOS if c in df.columns]
    ws.append(cols)
    for c_idx, _ in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for _, row in df.iterrows():
        ws.append([row.get(c, "") for c in cols])

    n_rows = ws.max_row
    n_cols = ws.max_column

    # Anchos y formato por columna
    for c_idx, col in enumerate(cols, 1):
        letter = get_column_letter(c_idx)
        ws.column_dimensions[letter].width = _COL_WIDTHS.get(col, 18)

    # Wrap text + alineacion por celda
    centrar = {"relevancia_1_5", "pais", "fecha_deteccion", "es_competidor",
               "plataforma", "estado", "categoria"}
    for r in range(2, n_rows + 1):
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(row=r, column=c_idx)
            cell.alignment = Alignment(
                horizontal="center" if col in centrar else "left",
                vertical="top",
                wrap_text=col in {"titulo_o_extracto", "senal_de_dolor",
                                  "mensaje_sugerido", "tema"},
            )
            # URL clicable
            if col == "URL" and cell.value:
                cell.hyperlink = str(cell.value)
                cell.font = Font(color="FF0563C1", underline="single")

    # Color de fila por relevancia (sobre toda la fila, no solo la celda)
    idx_rel = cols.index("relevancia_1_5") + 1
    idx_comp = cols.index("es_competidor") + 1 if "es_competidor" in cols else None
    for r in range(2, n_rows + 1):
        rel = ws.cell(row=r, column=idx_rel).value
        try:
            rel = int(rel)
        except Exception:
            rel = 1
        es_comp = False
        if idx_comp is not None:
            v = ws.cell(row=r, column=idx_comp).value
            es_comp = str(v).lower() in ("true", "1", "verdadero")
        color = _FILL_COMPETIDOR if es_comp else _FILL_REL.get(rel, _FILL_REL[1])
        fill = PatternFill("solid", fgColor=color)
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = fill

    # Estimar altura por fila segun el contenido envuelto mas largo.
    # Excel no hace autofit con wrap_text desde openpyxl, asi que lo hacemos a ojo:
    # lineas ≈ ceil(largo_texto / (ancho_columna * factor)) + saltos de linea explicitos.
    # Default font Calibri 11: ~1.1 chars por unidad de ancho de columna.
    WRAP_COLS = {"titulo_o_extracto", "senal_de_dolor", "mensaje_sugerido", "tema"}
    LINE_HEIGHT_PT = 15.0
    PADDING_PT = 6.0
    MIN_HEIGHT = 30.0
    MAX_HEIGHT = 420.0
    for r in range(2, n_rows + 1):
        max_lines = 1
        for c_idx, col in enumerate(cols, 1):
            if col not in WRAP_COLS:
                continue
            val = ws.cell(row=r, column=c_idx).value or ""
            text = str(val)
            width = _COL_WIDTHS.get(col, 18)
            chars_per_line = max(1, int(width * 1.1))
            explicit = text.count("\n") + 1
            wrap_lines = 0
            for segment in text.split("\n"):
                seg_len = len(segment)
                wrap_lines += max(1, -(-seg_len // chars_per_line))   # ceil
            lines = max(explicit, wrap_lines)
            if lines > max_lines:
                max_lines = lines
        height = max(MIN_HEIGHT, min(MAX_HEIGHT, max_lines * LINE_HEIGHT_PT + PADDING_PT))
        ws.row_dimensions[r].height = height

    # Panel fijo + autofiltro
    ws.freeze_panes = "B2"
    last_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A1:{last_letter}{n_rows}"

    # Dropdown en columna estado
    if "estado" in cols:
        idx_est = cols.index("estado") + 1
        letter_est = get_column_letter(idx_est)
        dv = DataValidation(type="list",
                            formula1='"' + ",".join(_ESTADOS) + '"',
                            allow_blank=True)
        dv.add(f"{letter_est}2:{letter_est}{n_rows}")
        ws.add_data_validation(dv)


def _escribir_excel(df: pd.DataFrame, xlsx_path: Path) -> None:
    wb = Workbook()
    # quitar la hoja por defecto
    default = wb.active
    wb.remove(default)

    # Hoja 1: prospectos accionables (relevancia >= 3, no competidor)
    df_top = df[(df["relevancia_1_5"].astype(int) >= 3) &
                (~df["es_competidor"].astype(bool))].copy()
    _hoja_tabla(wb, "Prospectos accionables", df_top)

    # Hoja 2: todo (ordenado)
    df_all = df.copy()
    _hoja_tabla(wb, "Todo", df_all)

    # Hoja 3: competidores aparte
    df_comp = df[df["es_competidor"].astype(bool)].copy()
    _hoja_tabla(wb, "Competidores", df_comp)

    # Hoja 4: descartados (relevancia 1-2, no competidor) para auditar
    df_ruido = df[(df["relevancia_1_5"].astype(int) <= 2) &
                  (~df["es_competidor"].astype(bool))].copy()
    _hoja_tabla(wb, "Ruido (rev. 1-2)", df_ruido)

    # Hoja 5: leyenda
    leyenda = wb.create_sheet("Leyenda")
    leyenda.append(["Color", "Significado"])
    for r, txt in [(5, "5 - cliente probable, accion inmediata"),
                   (4, "4 - dolor claro, actor LATAM o europeo claro"),
                   (3, "3 - senal interesante, requiere revisar"),
                   (2, "2 - senal debil"),
                   (1, "1 - ruido / descartar")]:
        leyenda.append(["", txt])
        cell = leyenda.cell(row=leyenda.max_row, column=1)
        cell.fill = PatternFill("solid", fgColor=_FILL_REL[r])
    leyenda.append(["", "competidor (cualquier relevancia)"])
    leyenda.cell(row=leyenda.max_row, column=1).fill = PatternFill(
        "solid", fgColor=_FILL_COMPETIDOR)
    leyenda.column_dimensions["A"].width = 10
    leyenda.column_dimensions["B"].width = 60

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def _escribir_markdown(df: pd.DataFrame, md_path: Path) -> None:
    lineas: list[str] = []
    lineas.append("# Seguimiento de prospectos - Synergium")
    lineas.append("")
    lineas.append(f"Generado: {date.today().isoformat()}. Total: {len(df)} registros.")
    lineas.append("")
    lineas.append("Ordenado por relevancia descendente.")
    lineas.append("")
    if df.empty:
        lineas.append("_Sin registros aun._")
        md_path.write_text("\n".join(lineas), encoding="utf-8")
        return

    cols_visibles = [
        "relevancia_1_5", "pais", "tipo_actor", "tema",
        "titulo_o_extracto", "estado", "URL",
    ]
    cab = "| " + " | ".join(cols_visibles) + " |"
    sep = "| " + " | ".join("---" for _ in cols_visibles) + " |"
    lineas.append(cab)
    lineas.append(sep)
    for _, row in df.iterrows():
        celdas = []
        for col in cols_visibles:
            val = str(row.get(col, ""))
            val = val.replace("|", "\\|").replace("\n", " ").strip()
            if col == "titulo_o_extracto" and len(val) > 120:
                val = val[:117] + "..."
            celdas.append(val)
        lineas.append("| " + " | ".join(celdas) + " |")

    # detalle de los top 10 con mensaje sugerido
    lineas.append("")
    lineas.append("## Top 10 con mensaje sugerido")
    lineas.append("")
    for _, row in df.head(10).iterrows():
        if not str(row.get("mensaje_sugerido", "")).strip():
            continue
        lineas.append(f"### [{row['relevancia_1_5']}] {row['titulo_o_extracto']}")
        lineas.append(f"- URL: {row['URL']}")
        lineas.append(f"- Pais: {row['pais']} - Actor: {row['tipo_actor']} - Tema: {row['tema']}")
        lineas.append(f"- Senial: {row['senal_de_dolor']}")
        lineas.append("")
        lineas.append("> " + str(row["mensaje_sugerido"]).replace("\n", "\n> "))
        lineas.append("")

    md_path.write_text("\n".join(lineas), encoding="utf-8")
